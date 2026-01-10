import pandas as pd
import random
from datetime import datetime
from uuid import uuid4
from typing import Dict, List, Optional
from fastapi import APIRouter, HTTPException

# Import internal modules
from .models import SessionInit, AnswerSubmit, SkipPhaseRequest
from .llm_service import real_time_correction

router = APIRouter()

# --- DATA INGESTION ---
try:
    VOCAB_DF = pd.read_csv("vocab.csv", sep=";")
    VOCAB_DF = VOCAB_DF.dropna(subset=['image_id', 'german_word'])
except Exception as e:
    print(f"Error loading vocab.csv: {e}")
    VOCAB_DF = pd.DataFrame()

# In-memory session cache
EXP_CACHE: Dict[str, dict] = {}

# --- UTILITY FUNCTIONS ---

def prepare_experiment_items():
    """Selects 20 unique words (4 per scenario) and sets items for Pre/Post tests."""
    scenarios = ["apartment_request", "travel", "swimming", "pet_sitting", "birthday"]
    learning_pool = []
    
    # Check if VOCAB_DF is empty or has insufficient data
    if VOCAB_DF.empty:
        print("[ERROR] VOCAB_DF is empty - cannot prepare experiment items")
        raise ValueError("No vocabulary data available. Please check vocab.csv")
    
    try:
        for sc in scenarios:
            cat_items = VOCAB_DF[VOCAB_DF['scenario'] == sc].to_dict('records')
            if len(cat_items) >= 4:
                learning_pool.extend(random.sample(cat_items, 4))
            else:
                # Use all available items for this scenario
                learning_pool.extend(cat_items)
                print(f"[WARNING] Scenario '{sc}' has only {len(cat_items)} items (expected 4)")

        # Ensure we have at least 5 items for test
        if len(learning_pool) < 5:
            print(f"[ERROR] Not enough items in learning pool: {len(learning_pool)} (need at least 5)")
            raise ValueError(f"Insufficient vocabulary data: only {len(learning_pool)} items found")

        test_items = random.sample(learning_pool, min(5, len(learning_pool)))
        
        task_types = ["article_mcq", "article_mcq", "plural_mcq", "plural_mcq", "type_word"]
        random.shuffle(task_types)
        
        pre_items = [item.copy() for item in test_items]
        post_items = [item.copy() for item in test_items]
        
        for i in range(len(pre_items)):
            pre_items[i]['assigned_task'] = task_types[i]
            post_items[i]['assigned_task'] = task_types[i]
            
        print(f"[INFO] Prepared experiment with {len(learning_pool)} learning items, {len(test_items)} test items")
        return learning_pool, pre_items, post_items
        
    except Exception as e:
        print(f"[ERROR] Failed to prepare experiment items: {e}")
        raise

# --- ENDPOINTS ---

@router.post("/experiment/init")
async def init_experiment(data: SessionInit):
    try:
        session_id = str(uuid4())
        learn_items, pre_items, post_items = prepare_experiment_items()
        
        EXP_CACHE[session_id] = {
            "user_id": data.user_id,
            "condition": data.condition,
            "phase": "pre-test",
            "current_index": 0,
            "attempt_count": 0,
            "items": {
                "pre-test": pre_items,
                "learning": learn_items,
                "post-test": post_items
            },
            "logs": []
        }
        
        print(f"[INFO] Initialized session {session_id} for user {data.user_id} (Condition {data.condition})")
        return {"session_id": session_id, "condition": data.condition}
        
    except ValueError as ve:
        # Data validation errors
        print(f"[ERROR] Validation error in init_experiment: {ve}")
        raise HTTPException(status_code=503, detail=f"Service unavailable: {str(ve)}")
    except Exception as e:
        # Unexpected errors
        print(f"[ERROR] Unexpected error in init_experiment: {type(e).__name__}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to initialize experiment: {str(e)}")

@router.get("/experiment/trial/{session_id}")
async def get_trial(session_id: str):
    session = EXP_CACHE.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found.")
    
    phase = session["phase"]
    idx = session["current_index"]
    
    if idx >= len(session["items"][phase]):
        if phase == "pre-test":
            session["phase"] = "learning"
            session["current_index"] = 0
        elif phase == "learning":
            session["phase"] = "post-test"
            session["current_index"] = 0
        else:
            return {"status": "completed"}
        
        phase = session["phase"]
        idx = 0

    item = session["items"][phase][idx]
    task_type = item.get('assigned_task', 'type_word')
    
    options = []
    if task_type == "article_mcq":
        options = ["der", "die", "das"]
    elif task_type == "plural_mcq":
        # FIXED: Distractors are now full words (noun + suffix) to match correct answer format
        correct = str(item['plural'])
        base = str(item['german_word'])
        
        # Possible plural patterns to generate distractors
        suffixes = ["en", "er", "e", "s", "n"]
        potential_distractors = [f"{base}{s}" for s in suffixes]
        
        # Filter out the correct one and select 2 random ones
        filtered_distractors = [d for d in potential_distractors if d.lower() != correct.lower()]
        options = list(set([correct] + random.sample(filtered_distractors, 2)))
        random.shuffle(options)

    image_url = f"/images/{item['image_id']}.jpg"

    return {
        "phase": phase,
        "index": idx,
        "total_in_phase": len(session["items"][phase]),
        "task_type": task_type,
        "image_url": image_url,
        "english_gloss": item['english_gloss'],
        "options": options if options else None,
        "german_word": item['german_word'] if task_type != "type_word" else None
    }

@router.post("/experiment/skip")
async def skip_to_phase(data: SkipPhaseRequest):
    """Jumps to a specific experiment phase for testing."""
    session = EXP_CACHE.get(data.session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if data.phase not in session["items"]:
        raise HTTPException(status_code=400, detail="Invalid phase")
    
    session["phase"] = data.phase
    session["current_index"] = 0
    session["attempt_count"] = 0
    return {"status": "ok", "new_phase": data.phase}

# ... imports
from .models import SessionInit, AnswerSubmit, SkipPhaseRequest, TutorRequest
from .eye_tracking_models import EyeTrackingFrame, BatchUploadRequest, FinalizeRequest
from .llm_service import real_time_correction, generate_tutor_response
import openpyxl
from openpyxl.utils import get_column_letter
from fastapi.responses import FileResponse
import tempfile
import os

# ...

# --- ENDPOINTS ---

@router.post("/experiment/tutor/ask")
async def ask_tutor(data: TutorRequest):
    """
    Direct endpoint for user questions to the tutor side-panel.
    """
    response = await generate_tutor_response(
        question=data.question,
        context=data.task_context.dict(),
        is_nudge=False,
        target_language=data.response_language
    )
    return response


@router.post("/experiment/init")
# ... (rest of init)

# ...

@router.post("/experiment/submit")
async def submit_answer(data: AnswerSubmit):
    session = EXP_CACHE.get(data.session_id)
    if not session:
        raise HTTPException(status_code=404)
    
    phase = session["phase"]
    idx = session["current_index"]
    item = session["items"][phase][idx]
    
    # Determine the correct answer based on the task type
    correct_val = f"{item['article']} {item['german_word']}"
    if item.get('assigned_task') == "article_mcq":
        correct_val = item['article']
    elif item.get('assigned_task') == "plural_mcq":
        correct_val = item['plural']

    is_correct = str(data.user_answer).strip() == str(correct_val).strip()
    
    # Tutor Nudge Logic (Condition B only - Adaptive)
    tutor_nudge = None
    if session["condition"] == "B":
        should_nudge = False
        if not is_correct:
            should_nudge = True # Always help on error
        else:
            # Deterministic praise: hash based (approx 20% of the time)
            combo_str = f"{item['german_word']}-{data.user_answer}"
            should_nudge = (hash(combo_str) % 5 == 0)

        if should_nudge:
            tutor_data = await generate_tutor_response(
                question="",
                context={
                    "prompt": item['english_gloss'],
                    "user_answer": data.user_answer,
                    "expected_answer": correct_val,
                    "is_correct": is_correct
                },
                is_nudge=True
            )
            tutor_nudge = tutor_data

    # 1. TEST PHASES (Pre-test / Post-test)
    if phase != "learning":
        session["current_index"] += 1
        return {
            "is_correct": is_correct, 
            "feedback": str(correct_val), 
            "move_next": True,
            "tutor_nudge": None # No AI help in tests
        }

    # 2. LEARNING PHASE
    if is_correct:
        session["current_index"] += 1
        session["attempt_count"] = 0
        return {
            "is_correct": True, 
            "feedback": f"Perfect! It is '{correct_val}'.", 
            "example": item['example_de'], 
            "move_next": True,
            "tutor_nudge": tutor_nudge
        }
    else:
        session["attempt_count"] += 1
        
        # Condition A: Static (Reveal immediately)
        if session["condition"] == "A":
            session["current_index"] += 1
            session["attempt_count"] = 0
            return {
                "is_correct": False, 
                "feedback": f"Incorrect. The correct answer is: {correct_val}.", 
                "move_next": True,
                "tutor_nudge": None
            }
        
        # Condition B: Adaptive (Hints then reveal)
        else:
            ai_data = None
            move_next = False
            
            if session["attempt_count"] >= 3:
                # Final Attempt: Reveal answer
                session["current_index"] += 1
                session["attempt_count"] = 0
                move_next = True
                
                # We use the new tutor logic as the primary feedback mechanism for the final reveal if available
                # But to preserve existing logic, we keep real_time_correction for the "feedback" field
                # and put the friendly tutor message in tutor_nudge
                
                ai_data = await real_time_correction(
                    data.user_answer, str(correct_val), 3, "A2", data.history
                )
            else:
                # Attempts 1 & 2
                ai_data = await real_time_correction(
                    data.user_answer, str(correct_val), session["attempt_count"], "A2", data.history
                )

            return {
                "is_correct": False, 
                "feedback": ai_data['tip'], 
                "move_next": move_next, 
                "example": item['example_de'] if move_next else None,
                "tutor_nudge": tutor_nudge
            }

@router.get("/experiment/export/{session_id}")
async def export_data(session_id: str):
    session = EXP_CACHE.get(session_id)
    if not session: raise HTTPException(status_code=404)
    return {
        "user_id": session["user_id"],
        "condition": session["condition"],
        "logs": session["logs"]
    }


# --- EYE-TRACKING ENDPOINTS ---

@router.post("/experiment/eye_tracking/batch/{session_id}")
async def upload_eye_tracking_batch(session_id: str, data: BatchUploadRequest):
    """
    Append a batch of eye-tracking frames to the session.
    Called every 3 minutes from the frontend during tracking.
    """
    session = EXP_CACHE.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Initialize eye_tracking structure if not exists
    if "eye_tracking" not in session:
        session["eye_tracking"] = {
            "active": True,
            "calibrated": True,
            "data": [],
            "finalized": False
        }
    
    # Append frames (convert Pydantic models to dicts)
    for frame in data.frames:
        session["eye_tracking"]["data"].append(frame.dict())
    
    frames_received = len(data.frames)
    total_frames = len(session["eye_tracking"]["data"])
    
    print(f"[INFO] Eye-tracking batch uploaded for session {session_id}: {frames_received} frames (total: {total_frames})")
    
    return {
        "status": "ok",
        "frames_received": frames_received,
        "total_frames": total_frames
    }


@router.post("/experiment/eye_tracking/finalize/{session_id}")
async def finalize_eye_tracking(session_id: str, data: FinalizeRequest):
    """
    Mark eye-tracking as finalized for this session.
    Called when user clicks "End tracking".
    """
    session = EXP_CACHE.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if "eye_tracking" not in session:
        raise HTTPException(status_code=400, detail="No eye-tracking data found for this session")
    
    session["eye_tracking"]["finalized"] = True
    session["eye_tracking"]["active"] = False
    
    total_frames = len(session["eye_tracking"]["data"])
    
    print(f"[INFO] Eye-tracking finalized for session {session_id}: {total_frames} total frames")
    
    return {
        "status": "ok",
        "total_frames": total_frames
    }


@router.get("/experiment/eye_tracking/export/{session_id}.xlsx")
async def export_eye_tracking_excel(session_id: str):
    """
    Generate and download Excel file with eye-tracking data.
    Exact column order as specified in requirements.
    """
    session = EXP_CACHE.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if "eye_tracking" not in session or not session["eye_tracking"]["data"]:
        raise HTTPException(status_code=400, detail="No eye-tracking data available for export")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "raw"
    
    # Define exact column order (14 columns)
    columns = [
        "timestamp_ms",
        "frame_idx",
        "phase",
        "screen_w_px",
        "screen_h_px",
        "face_detected",
        "left_iris_x_norm",
        "left_iris_y_norm",
        "right_iris_x_norm",
        "right_iris_y_norm",
        "gaze_x_px",
        "gaze_y_px",
        "gaze_valid",
        "on_screen"
    ]
    
    # Write header row
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data rows
    for row_idx, frame in enumerate(session["eye_tracking"]["data"], start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = frame.get(col_name)
            # Handle None/NaN values
            if value is None:
                value = float('nan')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()
    
    filename = f"eye_tracking_{session_id}.xlsx"
    
    print(f"[INFO] Eye-tracking Excel exported for session {session_id}: {len(session['eye_tracking']['data'])} frames")
    
    # Return file with proper headers
    return FileResponse(
        path=temp_file.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )